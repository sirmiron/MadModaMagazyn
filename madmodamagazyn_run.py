import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import openpyxl
import datetime
from datetime import date


class InventoryApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Analiza stanu magazynu")
        self.all_data = []

        # Frame for buttons (Polish UI)
        button_frame = tk.Frame(master)
        button_frame.pack(padx=10, pady=10, fill=tk.X)

        load_button = tk.Button(button_frame, text="Wybierz pliki i analizuj", command=self.load_files)
        load_button.pack(side=tk.LEFT, padx=5)

        save_button = tk.Button(button_frame, text="Zapisz do Excel", command=self.save_to_excel)
        save_button.pack(side=tk.LEFT, padx=5)

        # Frame for the detailed table (upper table)
        details_frame = tk.Frame(master)
        details_frame.pack(padx=10, pady=(10, 5), fill=tk.BOTH, expand=True)

        details_label = tk.Label(details_frame, text="Szczegóły")
        details_label.pack(anchor="w")

        # Detailed table columns: added "Typ / column B"
        self.details_columns = [
            "Towar", "Typ", "Index", "Cena zakupu",
            "Szt.", "Rozmiar", "Cena sprzedaży", "Plik"
        ]
        self.details_tree = ttk.Treeview(details_frame, columns=self.details_columns, show="headings")
        for col in self.details_columns:
            self.details_tree.heading(col, text=col)
            self.details_tree.column(col, width=100, anchor="w")
        self.details_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        details_scrollbar = tk.Scrollbar(details_frame, orient=tk.VERTICAL, command=self.details_tree.yview)
        self.details_tree.configure(yscrollcommand=details_scrollbar.set)
        details_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Frame for the summary table (lower table)
        summary_frame = tk.Frame(master)
        summary_frame.pack(padx=10, pady=(5, 5), fill=tk.BOTH, expand=True)

        summary_label = tk.Label(summary_frame, text="Podsumowanie (grupowanie po Towar, Typ, Index i Rozmiar)")
        summary_label.pack(anchor="w")

        # Summary table columns: added "Typ"
        self.summary_columns = [
            "Towar", "Typ", "Index", "Cena zakupu",
            "Szt.", "Rozmiar", "Cena sprzedaży", "Plik"
        ]
        self.summary_tree = ttk.Treeview(summary_frame, columns=self.summary_columns, show="headings")
        for col in self.summary_columns:
            self.summary_tree.heading(col, text=col)
            self.summary_tree.column(col, width=100, anchor="w")
        self.summary_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        summary_scrollbar = tk.Scrollbar(summary_frame, orient=tk.VERTICAL, command=self.summary_tree.yview)
        self.summary_tree.configure(yscrollcommand=summary_scrollbar.set)
        summary_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Frame for totals summary at the bottom
        totals_frame = tk.Frame(master)
        totals_frame.pack(padx=10, pady=(5, 10), fill=tk.X)
        self.totals_label = tk.Label(totals_frame, text="Podsumowanie całkowite: ")
        self.totals_label.pack(anchor="w")

    def display_error_table(self, error_list):
        """
        Displays a new window with a table of error details.
        The table contains columns: Plik, Wiersz, Komórka, Opis błędu, Wartość.
        If the error value is None, "Brak danych" is shown.
        Adds centered buttons "Zapisz błędy" and "Zamknij okno".
        """
        if not error_list:
            return  # Do not show window if no errors

        error_window = tk.Toplevel(self.master)
        error_window.title("Błędy importu danych")
        error_window.geometry("800x450")

        # Create a frame to hold the tree and scrollbar
        frame = tk.Frame(error_window)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        columns = ("Plik", "Wiersz", "Komórka", "Opis błędu", "Wartość")
        tree = ttk.Treeview(frame, columns=columns, show="headings")
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, anchor="w", width=150)
        tree.grid(row=0, column=0, sticky="nsew")

        v_scroll = tk.Scrollbar(frame, orient="vertical", command=tree.yview)
        v_scroll.grid(row=0, column=1, sticky="ns")
        tree.configure(yscrollcommand=v_scroll.set)

        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        for err in error_list:
            value = err.get("value") if err.get("value") is not None else "Brak danych"
            tree.insert("", tk.END, values=(
                err.get("file", ""),
                err.get("row", ""),
                err.get("col", ""),
                err.get("error", ""),
                value
            ))

        # Frame for action buttons
        btn_frame = tk.Frame(error_window)
        btn_frame.pack(fill=tk.X, pady=(0, 10), padx=10)
        btn_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(3, weight=1)

        save_btn = tk.Button(
            btn_frame,
            text="Zapisz błędy",
            command=lambda: self.save_errors_to_excel(error_list)
        )
        save_btn.grid(row=0, column=1, padx=5)

        close_btn = tk.Button(
            btn_frame,
            text="Zamknij okno",
            command=error_window.destroy
        )
        close_btn.grid(row=0, column=2, padx=5)

    def save_errors_to_excel(self, error_list):
        """
        Saves the list of error dicts to an Excel file.
        Filename defaults to 'YYYY-MM-DD_bledy_.xlsx'.
        """
        today_str = date.today().strftime("%Y-%m-%d")
        default_filename = f"{today_str}_bledy_.xlsx"

        save_path = filedialog.asksaveasfilename(
            title="Zapisz błędy do pliku Excel",
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Pliki Excel", "*.xlsx")]
        )
        if not save_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Błędy importu"

            headers = ["Plik", "Wiersz", "Komórka", "Opis błędu", "Wartość"]
            ws.append(headers)

            for err in error_list:
                value = err.get("value") if err.get("value") is not None else "Brak danych"
                ws.append([
                    err.get("file", ""),
                    err.get("row", ""),
                    err.get("col", ""),
                    err.get("error", ""),
                    value
                ])

            from openpyxl.utils import get_column_letter
            for col in ws.columns:
                max_len = max(len(str(cell.value)) for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

            wb.save(save_path)
            messagebox.showinfo("Sukces", f"Błędy zapisano do pliku:\n{save_path}")

        except Exception as e:
            messagebox.showerror("Błąd zapisu", f"Nie udało się zapisać pliku:\n{e}")

    def process_file(self, file_path):
        """
        Processes a single Excel file:
         - Reads the inventory date from cell G2 (formatted, but not used further).
         - Iterates over rows starting from row 5.
         - Adds a row only if:
             * The quantity ("Szt.") is greater than 0 and
             * The product name (column A) is not empty.
         - Additionally, logs an error if column B is empty.
         - Renames the column "Komis" to "Cena sprzedaży".
         - If the value in column "Index" (cell C) is not numeric, it is set to 0 and logged.
         - Converts "Index" to an integer.
         - Converts "Cena zakupu" and "Cena sprzedaży" to floats (rounded to 2 decimals).
         - Any conversion or missing‑value error in a row is recorded.
        """
        data_entries = []
        error_messages = []
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
        except Exception as e:
            messagebox.showerror(
                "Błąd",
                f"Nie udało się otworzyć pliku:\n{file_path}\n{e}"
            )
            return data_entries, error_messages

        inventory_date = ws['G2'].value
        if isinstance(inventory_date, (datetime.datetime, datetime.date)):
            inventory_date = inventory_date.strftime("%d-%m-%Y")

        for row in ws.iter_rows(min_row=5):
            row_num = row[0].row
            try:
                product_name = row[0].value      # column A
                col_b_value  = row[1].value      # column B
                quantity     = row[4].value      # column E
            except IndexError:
                continue

            if not (isinstance(quantity, (int, float)) and quantity > 0):
                continue
            if not (product_name and str(product_name).strip()):
                continue

            if not (col_b_value and str(col_b_value).strip()):
                error_messages.append({
                    "file": os.path.basename(file_path),
                    "row": row_num,
                    "col": row[1].coordinate,
                    "error": "Brak wartości w kolumnie B",
                    "value": col_b_value
                })

            try:
                index_val = int(float(row[2].value))
            except (ValueError, TypeError):
                error_messages.append({
                    "file": os.path.basename(file_path),
                    "row": row_num,
                    "col": row[2].coordinate,
                    "error": "Błąd konwersji 'Index'",
                    "value": row[2].value
                })
                index_val = 0

            try:
                price_purchase = round(float(row[3].value), 2)
            except (ValueError, TypeError):
                error_messages.append({
                    "file": os.path.basename(file_path),
                    "row": row_num,
                    "col": row[3].coordinate,
                    "error": "Błąd konwersji 'Cena zakupu'",
                    "value": row[3].value
                })
                price_purchase = 0.0

            try:
                price_sale = round(float(row[6].value), 2)
            except (ValueError, TypeError):
                error_messages.append({
                    "file": os.path.basename(file_path),
                    "row": row_num,
                    "col": row[6].coordinate,
                    "error": "Błąd konwersji 'Cena sprzedaży'",
                    "value": row[6].value
                })
                price_sale = 0.0

            entry = {
                "Towar": product_name,
                "Typ": col_b_value,
                "Index": index_val,
                "Cena zakupu": price_purchase,
                "Szt.": quantity,
                "Rozmiar": row[5].value,
                "Cena sprzedaży": price_sale,
                "Plik": os.path.basename(file_path)
            }
            data_entries.append(entry)

        return data_entries, error_messages

    def load_files(self):
        """
        Allows the user to select Excel files, processes each file,
        aggregates results and errors, then updates both tables.
        """
        file_paths = filedialog.askopenfilenames(
            title="Wybierz pliki Excel",
            filetypes=[("Pliki Excel", "*.xlsx")]
        )
        if not file_paths:
            return

        self.all_data = []
        all_errors = []
        for file_path in file_paths:
            entries, errors = self.process_file(file_path)
            self.all_data.extend(entries)
            all_errors.extend(errors)

        if all_errors:
            self.display_error_table(all_errors)

        if not self.all_data:
            messagebox.showinfo(
                "Informacja",
                "Nie znaleziono żadnych pozycji spełniających warunki."
            )
            return

        self.all_data.sort(key=lambda entry: entry["Index"])
        self.update_details_tree()
        self.update_summary_tree()

    def update_details_tree(self):
        """Updates the detailed table with data from all_data."""
        for row in self.details_tree.get_children():
            self.details_tree.delete(row)
        for entry in self.all_data:
            price_purchase_str = f"{entry['Cena zakupu']:.2f}"
            price_sale_str = f"{entry['Cena sprzedaży']:.2f}"
            self.details_tree.insert("", tk.END, values=(
                entry["Towar"],
                entry["Typ"],
                entry["Index"],
                price_purchase_str,
                entry["Szt."],
                entry["Rozmiar"],
                price_sale_str,
                entry["Plik"]
            ))

    def generate_summary(self):
        """
        Groups data by (Towar, Typ, Index, Rozmiar) and sums quantities and values.
        """
        summary = {}
        for entry in self.all_data:
            key = (entry["Towar"], entry["Typ"], entry["Index"], entry["Rozmiar"])
            if key not in summary:
                summary[key] = {
                    "Towar": entry["Towar"],
                    "Typ": entry["Typ"],
                    "Index": entry["Index"],
                    "Rozmiar": entry["Rozmiar"],
                    "Szt.": 0,
                    "Cena zakupu": 0.0,
                    "Cena sprzedaży": 0.0,
                    "Plik": set()
                }
            summary[key]["Szt."] += entry["Szt."]
            summary[key]["Cena zakupu"] += entry["Cena zakupu"]
            summary[key]["Cena sprzedaży"] += entry["Cena sprzedaży"]
            summary[key]["Plik"].add(entry["Plik"])

        result = []
        for data in summary.values():
            data["Plik"] = ", ".join(sorted(data["Plik"]))
            result.append(data)
        return result

    def update_summary_tree(self):
        """Updates the summary table with grouped data and updates the totals label."""
        try:
            for row in self.summary_tree.get_children():
                self.summary_tree.delete(row)

            summary_data = self.generate_summary()
            summary_data.sort(key=lambda x: (
                x["Towar"], str(x["Typ"]), x["Index"], x["Rozmiar"]
            ))

            total_quantity = 0
            total_purchase = 0.0
            total_sale = 0.0
            for entry in summary_data:
                price_purchase_str = f"{entry['Cena zakupu']:.2f}"
                price_sale_str = f"{entry['Cena sprzedaży']:.2f}"

                self.summary_tree.insert("", tk.END, values=(
                    entry["Towar"],
                    entry["Typ"],
                    entry["Index"],
                    price_purchase_str,
                    entry["Szt."],
                    entry["Rozmiar"],
                    price_sale_str,
                    entry["Plik"]
                ))

                total_quantity += entry["Szt."]
                total_purchase += entry["Cena zakupu"]
                total_sale += entry["Cena sprzedaży"]

            self.totals_label.config(
                text=(
                    f"Podsumowanie: Ilość = {total_quantity}, "
                    f"Wartość zakupu = {total_purchase:.2f}, "
                    f"Wartość sprzedaży = {total_sale:.2f}"
                )
            )
        except Exception as e:
            messagebox.showerror("Błąd", f"Nie udało się zaktualizować podsumowania:\n{e}")

    def adjust_column_widths(self, worksheet):
        """
        Adjusts column widths based on the maximum content length.
        """
        from openpyxl.utils import get_column_letter
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            worksheet.column_dimensions[col_letter].width = max_length + 2

    def save_to_excel(self):
        """
        Saves the data displayed in both tables to an Excel file.
        The first sheet ("Szczegóły") has detailed data,
        the second ("Suma") has the summary.
        Filename: 'stan_magazynu_YYYY-MM-DD.xlsx'.
        """
        if not self.details_tree.get_children():
            messagebox.showwarning("Brak danych", "Brak danych do zapisu. Najpierw przetwórz pliki.")
            return

        today_str = date.today().strftime("%Y-%m-%d")
        default_filename = f"stan_magazynu_{today_str}.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="Zapisz zbiorczą tabelę",
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Pliki Excel", "*.xlsx")]
        )
        if not save_path:
            return

        try:
            from openpyxl import Workbook
            wb = Workbook()

            # Szczegóły
            ws_details = wb.active
            ws_details.title = "Szczegóły"
            ws_details.append(self.details_columns)
            for child in self.details_tree.get_children():
                ws_details.append(self.details_tree.item(child)['values'])
            self.adjust_column_widths(ws_details)

            # Suma
            ws_summary = wb.create_sheet(title="Suma")
            ws_summary.append(self.summary_columns)
            summary_data = self.generate_summary()
            summary_data.sort(key=lambda x: (
                x["Towar"], str(x["Typ"]), x["Index"], x["Rozmiar"]
            ))
            for entry in summary_data:
                ws_summary.append([
                    entry["Towar"],
                    entry["Typ"],
                    entry["Index"],
                    f"{entry['Cena zakupu']:.2f}",
                    entry["Szt."],
                    entry["Rozmiar"],
                    f"{entry['Cena sprzedaży']:.2f}",
                    entry["Plik"]
                ])
            self.adjust_column_widths(ws_summary)

            wb.save(save_path)
            messagebox.showinfo("Sukces", f"Dane zostały zapisane do pliku:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Błąd", f"Nie udało się zapisać pliku:\n{e}")


if __name__ == "__main__":
    root = tk.Tk()
    root.iconbitmap("app_icon.ico")
    app = InventoryApp(root)
    root.mainloop()
